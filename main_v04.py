import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import openpyxl
from fpdf import FPDF

# Ruta del archivo Excel
file_path = "data_rutinas.xlsx"

# Preguntas para la encuesta general
questions = [
    ("Género:", ["Masculino", "Femenino"]),
    ("Tiempo por sesión:", ["Mayor a una hora", "Menor a una hora"]),
    ("Objetivo:", ["Aumento de masa muscular", "Aumento de fuerza", "Pérdida de grasa"]),
    ("¿Tienes limitaciones al levantar pesado?", ["Sí", "No"]),
    ("Edad:", ["Menor de 30 años", "Entre 30 y 50 años", "Mayor a 50 años"]),
    ("Nivel de entrenamiento:", ["Principiante", "Intermedio", "Avanzado"]),
    ("¿Tienes alguna lesión?", ["Sí", "No"]),
    ("Equipo disponible:", ["Mancuernas", "Peso corporal", "Gimnasio completo"]),
]

# Variables para datos personales y respuestas
personal_data = {}
user_answers = []
question_index = 0

# Mensaje de bienvenida
def show_welcome_message():
    messagebox.showinfo("Bienvenido", "¡Bienvenido al Sistema de Entrenamiento Personalizado!")

# Almacenar datos personales y abrir la encuesta
def store_personal_data():
    global personal_data
    try:
        name = name_entry.get()
        age = int(age_entry.get())
        height = float(height_entry.get())
        weight = float(weight_entry.get())

        if name and age > 0 and height > 0 and weight > 0:
            personal_data = {"Nombre": name, "Edad": age, "Estatura": height, "Peso": weight}
            personal_window.destroy()
            open_survey_window()
        else:
            messagebox.showerror("Error", "Por favor, ingresa valores válidos.")
    except ValueError:
        messagebox.showerror("Error", "Por favor, ingresa valores numéricos válidos.")

# Función para abrir la encuesta
def open_survey_window():
    global question_index, survey_window, label, dropdown
    show_welcome_message()

    question_index = 0
    user_answers.clear()

    survey_window = tk.Toplevel(root)
    survey_window.title("Encuesta General")
    survey_window.geometry("600x400")

    label = tk.Label(survey_window, text="Pregunta 1", font=("Arial", 14))
    label.pack(pady=20)

    dropdown = ttk.Combobox(survey_window, state="readonly")
    dropdown.pack(pady=10)

    tk.Button(survey_window, text="Siguiente", command=store_answer, bg="blue", fg="white").pack(pady=20)
    tk.Button(survey_window, text="Ver Historial", command=view_history).pack(pady=10)

    next_question()

# Función para mostrar la rutina seleccionada
def show_routine(sheet_name):
    try:
        routine_data = pd.read_excel(file_path, sheet_name=sheet_name)
        routine_window = tk.Toplevel(root)
        routine_window.title(f"Rutina: {sheet_name}")

        frame = tk.Frame(routine_window)
        frame.pack(pady=10, padx=10, fill="both", expand=True)

        tree = ttk.Treeview(frame, columns=list(routine_data.columns), show='headings', height=15)
        tree.pack(side="left", fill="both", expand=True)

        for col in routine_data.columns:
            tree.heading(col, text=col, anchor="center")
            tree.column(col, anchor="center", width=150)

        for _, row in routine_data.iterrows():
            tree.insert("", "end", values=list(row))

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la rutina: {e}")



# Función para guardar datos en Excel
def save_user_data(personal_data, answers, routine_sheet):
    try:
        workbook = openpyxl.load_workbook(file_path)

        # Si la hoja "usuarios" no existe, se crea
        if "usuarios" not in workbook.sheetnames:
            workbook.create_sheet("usuarios")

        sheet = workbook["usuarios"]

        # Si la hoja está vacía, agregar encabezados
        if sheet.max_row == 1 and not any(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers = [
                "Nombre", "Edad", "Estatura (cm)", "Peso (kg)", "Género",
                "Tiempo por sesión", "Objetivo", "Limitaciones al levantar pesado",
                "Rango de edad", "Nivel de entrenamiento", "Lesión",
                "Equipo disponible", "Rutina seleccionada"
            ]
            sheet.append(headers)

        # Preparar los datos del usuario
        user_data = [
            personal_data.get("Nombre", ""),
            personal_data.get("Edad", ""),
            personal_data.get("Estatura", ""),
            personal_data.get("Peso", ""),
            answers[0],  # Género
            answers[1],  # Tiempo por sesión
            answers[2] if answers[1] == "Mayor a una hora" else "No aplica",
            answers[3],  # Limitaciones
            answers[4],  # Rango de edad
            answers[5],  # Nivel
            answers[6],  # Lesión
            answers[7],  # Equipo
            routine_sheet if routine_sheet else "No encontrada"
        ]

        # Agregar los datos a la hoja
        sheet.append(user_data)

        # Guardar el archivo Excel
        workbook.save(file_path)
        print("Datos guardados correctamente en la hoja 'usuarios'.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron guardar los datos: {e}")


# Cargar la siguiente pregunta
def next_question():
    global question_index
    if question_index < len(questions):
        question, options = questions[question_index]
        label.config(text=question)
        dropdown["values"] = options
        dropdown.set(options[0])
        question_index += 1
    else:
        submit_answers()

# Guardar respuestas y avanzar en la encuesta
def store_answer():
    answer = dropdown.get()
    if answer:
        user_answers.append(answer)
        next_question()
    else:
        messagebox.showerror("Error", "Por favor, selecciona una opción antes de continuar.")

# Finalizar la encuesta y determinar la rutina personalizada
def submit_answers():
    if len(user_answers) == len(questions):
        routine_sheet = determine_routine(user_answers)
        save_user_data(personal_data, user_answers, routine_sheet)

        if routine_sheet:
            show_routine(routine_sheet)
        else:
            messagebox.showinfo("Sin Rutina", "No se encontró una rutina para las respuestas seleccionadas.")
    else:
        messagebox.showerror("Error", "Faltan respuestas en la encuesta.")

# Determinar la rutina personalizada
def determine_routine(answers):
    tiempo = answers[1]
    objetivo = answers[2] if tiempo == "Mayor a una hora" else None
    limitaciones = answers[3]
    edad = answers[4]
    nivel = answers[5]
    lesion = answers[6]
    equipo = answers[7]

    if equipo == "Peso corporal":
        return "Corporal"
    elif equipo == "Mancuernas":
        return "Mancuernas"
    elif equipo == "Gimnasio completo":
        if tiempo == "Menor a una hora":
            return "Rapida"
        elif limitaciones == "Sí":
            return "Peso_Bajo"
        elif edad == "Mayor a 50 años":
            return "Mayores"
        elif lesion == "Sí":
            return "Lesion"
        elif objetivo in ["Aumento de masa muscular", "Aumento de fuerza"]:
            if nivel == "Principiante":
                return "Masa_Principiante"
            elif nivel == "Intermedio":
                return "Masa_Intermedio"
            elif nivel == "Avanzado":
                return "Masa_Avanzado"
        elif objetivo == "Pérdida de grasa":
            if nivel == "Principiante":
                return "Perdida_Principiante"
            elif nivel == "Intermedio":
                return "Perdida_Intermedio"
            elif nivel == "Avanzado":
                return "Perdida_Avanzado"

    return None

# Exportar datos a PDF
def export_to_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(40, 10, "Datos del Usuario")

    pdf.set_font("Arial", "", 14)
    for key, value in personal_data.items():
        pdf.cell(0, 10, f"{key}: {value}", 0, 1)

    pdf.output("usuario_rutina.pdf")
    messagebox.showinfo("PDF Exportado", "Se exportaron los datos a 'usuario_rutina.pdf'.")

# Mostrar historial de datos almacenados
def view_history():
    try:
        history_data = pd.read_excel(file_path, sheet_name="usuarios")
        history_window = tk.Toplevel(root)
        history_window.title("Historial del Usuario")

        for index, row in history_data.iterrows():
            tk.Label(history_window, text=str(row)).pack()

    except Exception as e:
        messagebox.showerror("Error", f"No se encontró historial: {e}")

# Función para abrir la ventana de datos personales
def open_personal_data_window():
    global name_entry, age_entry, height_entry, weight_entry, personal_window
    personal_window = tk.Toplevel(root)
    personal_window.title("Datos Personales")
    personal_window.geometry("400x300")
    personal_window.configure(bg="lightblue")

    tk.Label(personal_window, text="Nombre:", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    name_entry = tk.Entry(personal_window)
    name_entry.pack(pady=5)

    tk.Label(personal_window, text="Edad:", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    age_entry = tk.Entry(personal_window)
    age_entry.pack(pady=5)

    tk.Label(personal_window, text="Estatura (cm):", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    height_entry = tk.Entry(personal_window)
    height_entry.pack(pady=5)

    tk.Label(personal_window, text="Peso (kg):", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    weight_entry = tk.Entry(personal_window)
    weight_entry.pack(pady=5)

    tk.Button(personal_window, text="Guardar", command=store_personal_data).pack(pady=20)

# Mostrar historial de datos almacenados y exportarlo a PDF
def view_history():
    try:
        history_data = pd.read_excel(file_path, sheet_name="usuarios")
        
        # Exportar a PDF
        export_history_to_pdf(history_data)

        # Mostrar historial estético en una ventana de Tkinter
        history_window = tk.Toplevel(root)
        history_window.title("Historial del Usuario")
        history_window.geometry("600x400")

        frame = tk.Frame(history_window)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        tree = ttk.Treeview(frame, columns=list(history_data.columns), show='headings')
        tree.pack(side="left", fill="both", expand=True)

        # Configuración de columnas y encabezados
        for col in history_data.columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=120)

        # Insertar filas en el Treeview
        for _, row in history_data.iterrows():
            tree.insert("", "end", values=list(row))

        # Scrollbar para el Treeview
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el historial: {e}")


# Exportar historial a PDF en orientación horizontal y en una sola hoja
def export_history_to_pdf(data):
    pdf = FPDF(orientation="L", unit="mm", format="A4")  # 'L' para orientación horizontal
    pdf.add_page()

    pdf.set_font("Arial", "B", 5)
    pdf.cell(0, 10, "Historial del Usuario", 0, 1, 'C')

    pdf.ln(5)
    pdf.set_font("Arial", "", 5)

    # Configuración de la anchura de columnas
    num_columns = len(data.columns)
    page_width = 250  # Ancho total de A4 en mm
    col_width = page_width / num_columns

    # Encabezados de columnas
    for column in data.columns:
        pdf.cell(col_width, 8, column, 1)
    pdf.ln()

    # Agregar filas de datos
    for index, row in data.iterrows():
        for item in row:
            pdf.cell(col_width, 10, str(item), 1)
        pdf.ln()

    # Guardar el archivo PDF
    pdf.output("historial_rutinas.pdf")
    messagebox.showinfo("PDF Exportado", "El historial se exportó a 'historial_rutinas.pdf'.")


# Crear la ventana principal
root = tk.Tk()
root.withdraw()

open_personal_data_window()

root.mainloop()
