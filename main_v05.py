import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from fpdf import FPDF
import openpyxl

# Ruta del archivo Excel
file_path = "data_rutinas.xlsx"

# Base de conocimiento: reglas para determinar la rutina personalizada
rules = [
    {"conditions": {"Equipo disponible": "Peso corporal"}, "result": "Corporal"},
    {"conditions": {"Equipo disponible": "Mancuernas"}, "result": "Mancuernas"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Menor a una hora", "¿Tienes limitaciones al levantar pesado?": "No"}, "result": "Rapida"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "¿Tienes limitaciones al levantar pesado?": "Sí"}, "result": "Peso_Bajo"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Edad": "Mayor a 50 años"}, "result": "Mayores"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "¿Tienes alguna lesión?": "Sí"}, "result": "Lesion"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Aumento de masa muscular", "Nivel de entrenamiento": "Principiante", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Masa_Principiante"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Aumento de masa muscular", "Nivel de entrenamiento": "Intermedio", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Masa_Intermedio"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Aumento de masa muscular", "Nivel de entrenamiento": "Avanzado", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Masa_Avanzado"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Pérdida de grasa", "Nivel de entrenamiento": "Principiante", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Perdida_Principiante"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Pérdida de grasa", "Nivel de entrenamiento": "Intermedio", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Perdida_Intermedio"},
    {"conditions": {"Equipo disponible": "Gimnasio completo", "Tiempo por sesión": "Mayor a una hora", "Objetivo": "Pérdida de grasa", "Nivel de entrenamiento": "Avanzado", "¿Tienes limitaciones al levantar pesado?": "No", "¿Tienes alguna lesión?": "No", "Edad": "Menor de 30 años"}, "result": "Perdida_Avanzado"}
]

# Función del motor de inferencia para determinar la rutina personalizada
def infer_routine(answers):
    print("Respuestas del usuario:", answers)  # Registro de las respuestas del usuario para depuración
    for idx, rule in enumerate(rules):
        print(f"Evaluando regla {idx + 1}: {rule['conditions']}")
        conditions_matched = [key for key, value in rule["conditions"].items() if answers.get(key) == value]
        conditions_failed = [key for key, value in rule["conditions"].items() if answers.get(key) != value]

        print(f"Condiciones que coinciden: {conditions_matched}")
        print(f"Condiciones que fallaron: {conditions_failed}")

        if all(answers.get(key) == value for key, value in rule["conditions"].items()):
            print(f"Regla {idx + 1} aplicada: {rule['result']}")
            return rule["result"]
    print("Ninguna regla fue aplicada.")
    return None

# Preguntas para la encuesta general
questions = [
    ("Género:", ["Masculino", "Femenino"]),
    ("Tiempo por sesión:", ["Mayor a una hora", "Menor a una hora"]),
    ("Objetivo:", ["Aumento de masa muscular", "Aumento de fuerza", "Pérdida de grasa"]),
    ("¿Tienes limitaciones al levantar pesado?", ["Sí", "No"]),
    ("Edad:", ["Menor de 30 años", "Entre 30 y 50 años", "Mayor a 50 años"]),
    ("Nivel de entrenamiento:", ["Principiante", "Intermedio", "Avanzado"]),
    ("¿Tienes alguna lesión?", ["Sí", "No"]),
    ("Equipo disponible:", ["Mancuernas", "Peso corporal", "Gimnasio completo"]),
]

# Variables para datos personales y respuestas
personal_data = {}
user_answers = {}
question_index = 0

# Mensaje de bienvenida
def show_welcome_message():
    messagebox.showinfo("Bienvenido", "¡Bienvenido al Sistema de Entrenamiento Personalizado!")

# Almacenar datos personales y abrir la encuesta
def store_personal_data():
    global personal_data
    try:
        name = name_entry.get()
        age = int(age_entry.get())
        height = float(height_entry.get())
        weight = float(weight_entry.get())

        if name and age > 0 and height > 0 and weight > 0:
            personal_data = {"Nombre": name, "Edad": age, "Estatura": height, "Peso": weight}
            personal_window.destroy()
            open_survey_window()
        else:
            messagebox.showerror("Error", "Por favor, ingresa valores válidos.")
    except ValueError:
        messagebox.showerror("Error", "Por favor, ingresa valores numéricos válidos.")

# Función para abrir la encuesta
def open_survey_window():
    global question_index, survey_window, label, dropdown
    show_welcome_message()

    question_index = 0
    user_answers.clear()

    survey_window = tk.Toplevel(root)
    survey_window.title("Encuesta General")
    survey_window.geometry("600x400")

    label = tk.Label(survey_window, text="Pregunta 1", font=("Arial", 14))
    label.pack(pady=20)

    dropdown = ttk.Combobox(survey_window, state="readonly")
    dropdown.pack(pady=10)

    tk.Button(survey_window, text="Siguiente", command=store_answer, bg="blue", fg="white").pack(pady=20)

    next_question()

# Mostrar la rutina seleccionada
def show_routine(sheet_name):
    try:
        routine_data = pd.read_excel(file_path, sheet_name=sheet_name)
        routine_window = tk.Toplevel(root)
        routine_window.title(f"Rutina: {sheet_name}")

        frame = tk.Frame(routine_window)
        frame.pack(pady=10, padx=10, fill="both", expand=True)

        tree = ttk.Treeview(frame, columns=list(routine_data.columns), show='headings', height=15)
        tree.pack(side="left", fill="both", expand=True)

        for col in routine_data.columns:
            tree.heading(col, text=col, anchor="center")
            tree.column(col, anchor="center", width=150)

        for _, row in routine_data.iterrows():
            tree.insert("", "end", values=list(row))

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la rutina: {e}")

# Guardar datos del usuario en Excel
def save_user_data(personal_data, answers, routine_sheet):
    try:
        workbook = openpyxl.load_workbook(file_path)

        if "usuarios" not in workbook.sheetnames:
            workbook.create_sheet("usuarios")

        sheet = workbook["usuarios"]

        if sheet.max_row == 1 and not any(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers = ["Nombre", "Edad", "Estatura (cm)", "Peso (kg)"] + list(answers.keys()) + ["Rutina seleccionada"]
            sheet.append(headers)

        user_data = [
            personal_data.get("Nombre", ""),
            personal_data.get("Edad", ""),
            personal_data.get("Estatura", ""),
            personal_data.get("Peso", ""),
        ] + list(answers.values()) + [routine_sheet]

        sheet.append(user_data)
        workbook.save(file_path)
        print("Datos guardados correctamente en la hoja 'usuarios'.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron guardar los datos: {e}")

# Cargar la siguiente pregunta
def next_question():
    global question_index
    if question_index < len(questions):
        question, options = questions[question_index]
        label.config(text=question)
        dropdown["values"] = options
        dropdown.set(options[0])
        question_index += 1
    else:
        submit_answers()

# Guardar respuestas y avanzar en la encuesta
def store_answer():
    answer = dropdown.get()
    if answer:
        key = questions[question_index - 1][0]
        user_answers[key] = answer
        next_question()
    else:
        messagebox.showerror("Error", "Por favor, selecciona una opción antes de continuar.")

# Finalizar la encuesta y determinar la rutina personalizada
def submit_answers():
    if len(user_answers) == len(questions):
        routine_sheet = infer_routine(user_answers)
        save_user_data(personal_data, user_answers, routine_sheet)

        if routine_sheet:
            show_routine(routine_sheet)
        else:
            messagebox.showinfo("Sin Rutina", "No se encontró una rutina para las respuestas seleccionadas.")
    else:
        messagebox.showerror("Error", "Faltan respuestas en la encuesta.")

# Crear la ventana para datos personales
def open_personal_data_window():
    global name_entry, age_entry, height_entry, weight_entry, personal_window
    personal_window = tk.Toplevel(root)
    personal_window.title("Datos Personales")
    personal_window.geometry("400x300")
    personal_window.configure(bg="lightblue")

    tk.Label(personal_window, text="Nombre:", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    name_entry = tk.Entry(personal_window)
    name_entry.pack(pady=5)

    tk.Label(personal_window, text="Edad:", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    age_entry = tk.Entry(personal_window)
    age_entry.pack(pady=5)

    tk.Label(personal_window, text="Estatura (cm):", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    height_entry = tk.Entry(personal_window)
    height_entry.pack(pady=5)

    tk.Label(personal_window, text="Peso (kg):", font=('Helvetica', 12), bg="lightblue").pack(pady=5)
    weight_entry = tk.Entry(personal_window)
    weight_entry.pack(pady=5)

    tk.Button(personal_window, text="Guardar", command=store_personal_data).pack(pady=20)

# Crear la ventana principal
root = tk.Tk()
root.withdraw()

open_personal_data_window()

root.mainloop()
